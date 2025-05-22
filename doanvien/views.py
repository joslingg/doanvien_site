from django.shortcuts import render,redirect, get_object_or_404
from .forms import MemberForm
from django.http import HttpResponse
from django.db.models import Q
from .models import Member,Department,Group
from django.db.models import Count
import openpyxl
from openpyxl.styles import Font
import pandas as pd

def home(request):
    return HttpResponse("<h2>Chào bạn đến với hệ thống quản lý Đoàn viên!</h2>")

def add_member(request):
    if request.method =='POST':
        form = MemberForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('member-list')
    else:
         form = MemberForm()
    return render(request, 'doanvien/add_member.html',{'form':form})

def member_list(request):
    search = request.GET.get('q','')
    department_filter = request.GET.get('department','')
    group_filter = request.GET.get('group','')
    status_filter = request.GET.get('status','')
    note_filter = request.GET.get('note','')
    
    sort = request.GET.get('sort','name')
    direction = request.GET.get('direction','asc')
    
    members = Member.objects.all()
    if search:
        members = members.filter(name__icontains=search)
    
    if department_filter:
        members = members.filter(department_id=department_filter)
        
    if group_filter:
        members = members.filter(group_id=group_filter)
        
    if status_filter:
        members = members.filter(status = status_filter)
        
    if note_filter:
        members = members.filter(note = note_filter)
    
    if direction == 'desc':
        members = members.order_by(f'-{sort}')
    else:
        members = members.order_by(sort)
        
    departments = Department.objects.all()
    groups = Group.objects.all()
    
    status_choices = [choice[0] for choice in Member._meta.get_field('status').choices]
    note_choices = [choice[0] for choice in Member._meta.get_field('note').choices]
       
    return render(request, 'doanvien/member_list.html',{
        'members':members,
        'search': search,
        'departments': departments,
        'groups': groups,
        'selected_department': department_filter,
        'selected_group': group_filter,
        'selected_status': status_filter,
        'statuses': status_choices,
        'selected_note': note_filter,
        'notes': note_choices,
        'query_string': request.GET.urlencode(),
        'current_sort': sort,
        'current_direction': direction,
    })
    
def edit_member(request,pk):
    member = get_object_or_404(Member,pk=pk)
    if request.method == 'POST':
        form = MemberForm(request.POST, instance=member)
        if form.is_valid():
            form.save()
            return redirect('member-list')
        else:   
            return render(request, 'doanvien/edit_member.html',{'form':form,'member':member})
    else:
        form = MemberForm(instance=member)
        
    return render(request, 'doanvien/edit_member.html', {'form': form, 'member': member})

def delete_member(request,pk):
    member = get_object_or_404(Member,pk=pk)
    if request.method == 'POST':
        member.delete()
        return redirect('member-list')

    return render(request,'doanvien/delete_member.html',{'member':member})

def member_stats(request):
    department_stats = (
        Member.objects.values('department__name')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    
    group_stats = (
        Member.objects.values('group__name')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    
    return render(request, 'doanvien/member_stats.html', {
        'department_stats': department_stats,
        'group_stats': group_stats,
    })
    
def export_member_list(request):
    # Tạo Workbook mới
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách Đoàn viên"

    # Header
    headers = ['STT', 'Họ tên', 'Giới tính', 'Khoa/Phòng', 'Tổ', 'Tình trạng', 'Ghi chú']
    ws.append(headers)

    # In đậm Header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Lấy dữ liệu từ database
    members = Member.objects.all()
    for index, member in enumerate(members,start=1):
        ws.append([
            index,
            member.name,
            member.gender,
            member.department.name if member.department else "",
            member.group.name if member.group else "",
            member.status,
            member.note if member.note else ""
        ])

    # Tạo response để tải file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Danh_sach_doan_vien.xlsx'
    wb.save(response)

    return response

from .models import Department, Group, Member

def import_member_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['file']
        df = pd.read_excel(excel_file)
        
        for _, row in df.iterrows():
            # Lấy đối tượng Department, nếu chưa có thì tạo mới
            department, _ = Department.objects.get_or_create(name=row['KHOA'])
            # Lấy đối tượng Group, nếu chưa có thì tạo mới
            group, _ = Group.objects.get_or_create(name=row['TO'])
            
            # Kiểm tra giá trị của "NOTE" có khớp với choices hay không
            note_value = row['GHI_CHU'] if row['GHI_CHU'] in dict(Member._meta.get_field('note').choices) else None
            
            # Tạo mới Member
            Member.objects.create(
                name=row['TEN'],
                gender=row['GIOI_TINH'],
                department=department,   # Gán ForeignKey
                group=group,             # Gán ForeignKey
                status=row['TINH_TRANG'],
                note=note_value          # Gán giá trị note
            )
        return redirect('member-list')
    return render(request, 'doanvien/import_members.html')

def download_template(quest):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Temlate Import"
    
    headers = ['TEN', 'GIOI_TINH', 'KHOA', 'TO', 'TINH_TRANG', 'GHI_CHU']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=import_template.xlsx'
    wb.save(response)
    return response
